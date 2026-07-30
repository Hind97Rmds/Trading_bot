"""
backtest.py — Idealized Gann backtest engine.

Owns:
  - run_gann_backtest (zero-friction historical engine)
  - BtProgress (Telegram progress bar)
  - _build_gann_cycle_defs (cycle-anchoring logic)
  - Optional MT5-exported CSV data source
  - Excel export formatting
"""

import asyncio
import os
import time
from datetime import datetime, timedelta, timezone

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from state import (
    bot_state, SYMBOL_INFO, _TFS, DAM_OFF, _safe_float, log_exception, c_log,
    _safe_task, get_http, TG_TOKEN,
)
from market_data import fetch_candles
from strategy import (
    gann_calc_levels, gann_active_levels, _gann_atr, _gann_tf_tp,
    _gann_tf_sl, core_eval_break_even, _anchor_hours, _anchor_label,
    _last_closed_anchor_time_utc, _is_within_dam_restricted_window,
    _DAM_RESTRICTED_WINDOWS,
)
from state import DAM_OFF


# ── Gann Dynamic Recalc Cadence ──
GANN_DYNAMIC_RECALC_MINUTES = 5


# ═══════════════════════════════════════════════════════════════════
# MT5-EXPORTED CSV DATA SOURCE (optional, opt-in alternative to OANDA)
# ═══════════════════════════════════════════════════════════════════
# Why this exists: run_gann_backtest normally pulls every candle from OANDA
# (via fetch_candles). The MQL5 EA/Strategy Tester instead uses the broker's
# OWN price history. Two different vendors will occasionally disagree on an
# H1/M1 candle's exact OHLC by a few cents -- and because every Gann level is
# a percentage offset of one anchor close, that tiny disagreement cascades
# into a different level ladder, which can flip whether a level gets touched
# at all. There is no code fix for "two vendors recorded the market
# slightly differently" -- the only real fix is feeding BOTH systems from
# the exact same history. This section lets the Python backtest read
# broker-exported CSV files instead of calling OANDA, when they're present.
#
# Expected file naming (place these in bot_state['csv_data_dir']):
#   <SYMBOL>_<TF>.csv   e.g.  XAUUSD_H1.csv , XAUUSD_M1.csv , XAUUSD_M5.csv
# (underscores in the bot's own symbol id like "XAU_USD" are stripped when
# building this filename, to match MT5's own symbol naming convention.)
#
# Expected columns (matches MT5's Symbols > Bars > Export dialog output,
# comma- or tab-delimited, with or without the <ANGLE_BRACKET> header style):
#   DATE, TIME, OPEN, HIGH, LOW, CLOSE, [TICKVOL], [VOL], [SPREAD]
# DATE/TIME are in the BROKER SERVER's own local time, not UTC -- set
# bot_state['csv_broker_utc_offset_hours'] to that server's UTC offset
# (e.g. 3 for a server running UTC+3) so timestamps convert correctly and
# line up with OANDA's UTC-based candles for a fair comparison.

_TF_TO_MT5_SUFFIX = {
    '1m': 'M1', '2m': 'M2', '3m': 'M3', '4m': 'M4', '5m': 'M5', '6m': 'M6',
    '10m': 'M10', '15m': 'M15', '20m': 'M20', '30m': 'M30', '1h': 'H1', '2h': 'H2', '4h': 'H4',
}
_csv_candle_cache: dict[str, list] = {}  # file path -> parsed candles, loaded once per process


def _load_csv_candle_file(csv_path: str, broker_utc_offset_hours: float) -> list:
    if csv_path in _csv_candle_cache:
        return _csv_candle_cache[csv_path]
    try:
        df = pd.read_csv(csv_path, sep=None, engine='python')
        df.columns = [str(c).strip().strip('<>').upper() for c in df.columns]
        needed = ['DATE', 'TIME', 'OPEN', 'HIGH', 'LOW', 'CLOSE']
        missing = [c for c in needed if c not in df.columns]
        if missing:
            c_log(f"CSV import [{csv_path}]: missing required columns {missing}, columns found: {list(df.columns)}")
            _csv_candle_cache[csv_path] = []
            return []
        vol_col = 'TICKVOL' if 'TICKVOL' in df.columns else ('VOL' if 'VOL' in df.columns else None)

        naive = pd.to_datetime(df['DATE'].astype(str) + ' ' + df['TIME'].astype(str),
                                errors='coerce')
        utc_time = naive - pd.Timedelta(hours=broker_utc_offset_hours)
        utc_time = utc_time.dt.tz_localize('UTC')

        out = []
        for i in range(len(df)):
            if pd.isna(naive.iloc[i]):
                continue
            out.append({
                'time': utc_time.iloc[i],
                'open': float(df['OPEN'].iloc[i]), 'high': float(df['HIGH'].iloc[i]),
                'low': float(df['LOW'].iloc[i]), 'close': float(df['CLOSE'].iloc[i]),
                'volume': float(df[vol_col].iloc[i]) if vol_col else 1.0,
            })
        out.sort(key=lambda c: c['time'])
        c_log(f"CSV import [{csv_path}]: loaded {len(out)} candles "
              f"({out[0]['time']} → {out[-1]['time']})" if out else f"CSV import [{csv_path}]: 0 usable rows")
        _csv_candle_cache[csv_path] = out
        return out
    except Exception as e:
        log_exception(f'_load_csv_candle_file [{csv_path}]', e)
        _csv_candle_cache[csv_path] = []
        return []


# Tracks what the LAST run of run_gann_backtest
# actually used, per (symbol, granularity) -- reset at the start of each
# run via _reset_data_source_stats(), read back at the end to report the
# true source used, not just the configured intention.
_data_source_stats = {'csv_hits': {}, 'oanda_fallbacks': {}, 'csv_files_missing': set()}


def _reset_data_source_stats():
    _data_source_stats['csv_hits'] = {}
    _data_source_stats['oanda_fallbacks'] = {}
    _data_source_stats['csv_files_missing'] = set()


def _format_data_source_report() -> str:
    csv_dir = bot_state.get('csv_data_dir')
    if not csv_dir:
        return "📡 مصدر البيانات: OANDA (وضع CSV غير مفعّل)"
    hits = _data_source_stats['csv_hits']
    misses = _data_source_stats['oanda_fallbacks']
    missing_files = _data_source_stats['csv_files_missing']
    lines = [f"📂 مجلد CSV المُعتمَد: <code>{csv_dir}</code>"]
    if hits:
        hit_str = ', '.join(f"{tf}: {n} شمعة" for tf, n in sorted(hits.items()))
        lines.append(f"✅ قُرئ فعلياً من CSV → {hit_str}")
    if misses:
        miss_str = ', '.join(f"{tf}: {n} استدعاء" for tf, n in sorted(misses.items()))
        lines.append(f"☁️ رجع لـOANDA فعلياً (لم يجد ملف CSV مطابق) → {miss_str}")
    if missing_files:
        lines.append("🔍 أسماء الملفات التي بحث عنها ولم يجدها:\n" + '\n'.join(f"  • <code>{f}</code>" for f in sorted(missing_files)))
    if not hits and not misses:
        lines.append("⚠️ لم يُستدعَ أي جلب بيانات بعد.")
    return '\n'.join(lines)


async def _fetch_candles_hybrid(symbol: str, granularity_str: str, count: int = 5000,
                                 end_time: datetime = None) -> list:
    """Drop-in replacement for fetch_candles(): uses a matching MT5-exported
    CSV file when bot_state['csv_data_dir'] is set AND the file exists,
    otherwise transparently falls back to the normal OANDA fetch_candles().
    Records which source was actually used in _data_source_stats so the
    caller can report ground truth, not just the configured intention."""
    csv_dir = bot_state.get('csv_data_dir')
    if csv_dir:
        suffix = _TF_TO_MT5_SUFFIX.get(granularity_str)
        if suffix:
            fname = f"{symbol.replace('_', '')}_{suffix}.csv"
            path = os.path.join(csv_dir, fname)
            if os.path.isfile(path):
                offset = bot_state.get('csv_broker_utc_offset_hours', 3.0)
                candles = _load_csv_candle_file(path, offset)
                if candles:
                    end = end_time if end_time else datetime.now(timezone.utc)
                    end_ts = pd.Timestamp(end).tz_convert('UTC') if pd.Timestamp(end).tzinfo else pd.Timestamp(end, tz='UTC')
                    filtered = [c for c in candles if c['time'] <= end_ts]
                    result = filtered[-count:] if count else filtered
                    _data_source_stats['csv_hits'][granularity_str] = \
                        _data_source_stats['csv_hits'].get(granularity_str, 0) + len(result)
                    return result
            else:
                _data_source_stats['csv_files_missing'].add(fname)
        _data_source_stats['oanda_fallbacks'][granularity_str] = \
            _data_source_stats['oanda_fallbacks'].get(granularity_str, 0) + 1
    return await fetch_candles(symbol, granularity_str, count=count, end_time=end_time)


def _build_gann_cycle_defs(sym_state: dict, valid_h1: list, mc_1m: list) -> list[dict]:
    mode = bot_state.get('gann_calculation_mode', 'static_h1')
    cycle_h = sym_state['gann_cycle_hours']
    if mode != 'dynamic_live':
        out = []
        for h1 in valid_h1:
            t_start = h1['time'] + timedelta(hours=1)
            out.append({'t_start': t_start, 't_end': t_start + timedelta(hours=cycle_h),
                         'close': float(h1['close'])})
        return out
    px_series = sorted(mc_1m, key=lambda c: c['time']) if mc_1m else []
    if not px_series:
        return []
    out = []
    bucket = px_series[0]['time'].floor(f'{GANN_DYNAMIC_RECALC_MINUTES}min')
    last_t = px_series[-1]['time']
    i = 0; n = len(px_series)
    while bucket <= last_t:
        while i + 1 < n and px_series[i + 1]['time'] <= bucket:
            i += 1
        if px_series[i]['time'] <= bucket:
            out.append({'t_start': bucket,
                         't_end': bucket + timedelta(minutes=GANN_DYNAMIC_RECALC_MINUTES),
                         'close': float(px_series[i]['close'])})
        bucket += timedelta(minutes=GANN_DYNAMIC_RECALC_MINUTES)
    return out


# ── Backtest Progress Tracker ──
class BtProgress:
    BAR_LEN = 14; HEARTBEAT = 15

    def __init__(self, label: str, active_tfs: list):
        self.label = label; self.active_tfs = active_tfs; self.cancelled = False
        self.phase = 'Initialising...'; self.tf_done = 0; self.tf_total = len(active_tfs)
        self.current_tf = ''; self.bars_done = 0; self.bars_total = 0
        self.win = 0; self.loss = 0; self.be = 0; self.profit = 0.0
        self.chat_id = None; self.msg_id = None; self._last_edit = 0.0
        self._lock = asyncio.Lock(); self._hb_task = None; self._start_ts = 0.0

    def _bar(self, done: int, total: int) -> str:
        if total == 0: return chr(9617) * self.BAR_LEN
        filled = round(done / total * self.BAR_LEN)
        return chr(9608) * filled + chr(9617) * (self.BAR_LEN - filled)

    def _elapsed(self) -> str:
        secs = int(datetime.now(timezone.utc).timestamp() - self._start_ts)
        m, s = divmod(secs, 60); return f'{m}m {s:02d}s'

    def _build_text(self) -> str:
        total = self.win + self.loss
        wr = f'{round(self.win / total * 100)}%' if total else '-'
        pnl = f'+${round(self.profit,2)}' if self.profit >= 0 else f'-${abs(round(self.profit,2))}'
        icon = '▲' if self.profit >= 0 else '▼'
        overall = (self.tf_done + self.bars_done / self.bars_total) / max(self.tf_total, 1) if self.bars_total else self.tf_done / max(self.tf_total, 1)
        ov_bar = self._bar(round(overall * 100), 100); ov_pct = f'{round(overall * 100)}%'
        tf_bar = self._bar(self.bars_done, self.bars_total) if self.bars_total else chr(9617) * self.BAR_LEN
        tf_pct = f'{round(self.bars_done / self.bars_total * 100)}%' if self.bars_total else '-'
        lines = [f'Backtest — <b>{self.label}</b>', f'<b>Phase:</b> {self.phase}', '',
                 f'<b>Overall</b>  {ov_pct}', f'<code>[{ov_bar}]</code>']
        if self.current_tf:
            lines += ['', f'<b>TF:</b> {self.current_tf}  ({self.tf_done}/{self.tf_total})',
                      f'<code>[{tf_bar}] {tf_pct}</code>', f'Bars: {self.bars_done}/{self.bars_total}']
        lines += ['', f'W:{self.win}  L:{self.loss}  BE:{self.be}', f'{icon} {pnl}  WR:{wr}', '',
                  f'Elapsed: {self._elapsed()}']
        if self.cancelled: lines.append('<b>CANCELLED</b>')
        return '\n'.join(lines)

    async def start(self, chat_id: int) -> None:
        self.chat_id = chat_id
        self._start_ts = datetime.now(timezone.utc).timestamp()
        self._last_edit = self._start_ts
        payload = {'chat_id': chat_id, 'text': self._build_text(), 'parse_mode': 'HTML',
                   'reply_markup': {'inline_keyboard': [[{'text': '⏹ Cancel', 'callback_data': 'cancel_bt'}]]}}
        try:
            import aiohttp
            from state import TG_TOKEN
            async with aiohttp.ClientSession() as sess:
                async with sess.post(f'https://api.telegram.org/bot{TG_TOKEN}/sendMessage', json=payload) as resp:
                    if resp.status == 200:
                        self.msg_id = (await resp.json())['result']['message_id']
        except Exception:
            pass
        self._hb_task = asyncio.create_task(self._heartbeat())

    async def _heartbeat(self) -> None:
        while not self.cancelled:
            await asyncio.sleep(self.HEARTBEAT)
            await self._edit(force=True)

    async def _edit(self, force: bool = False) -> None:
        now = datetime.now(timezone.utc).timestamp()
        if not force and (now - self._last_edit) < 3:
            return
        if not self.msg_id or not self.chat_id:
            return
        async with self._lock:
            self._last_edit = now
            import aiohttp
            from state import TG_TOKEN
            payload = {'chat_id': self.chat_id, 'message_id': self.msg_id,
                       'text': self._build_text(), 'parse_mode': 'HTML'}
            if not self.cancelled:
                payload['reply_markup'] = {'inline_keyboard': [[{'text': '⏹ Cancel', 'callback_data': 'cancel_bt'}]]}
            try:
                async with aiohttp.ClientSession() as sess:
                    await sess.post(f'https://api.telegram.org/bot{TG_TOKEN}/editMessageText', json=payload)
            except Exception:
                pass

    async def set_phase(self, phase: str) -> None: self.phase = phase; await self._edit()
    async def set_tf(self, tf: str, bars_total: int) -> None:
        self.current_tf = tf; self.bars_done = 0; self.bars_total = bars_total; await self._edit(force=True)
    async def tick(self, bar_n: int, win: int, loss: int, be: int, profit: float) -> None:
        self.bars_done = bar_n; self.win = win; self.loss = loss; self.be = be
        self.profit = profit; await self._edit()
    async def done(self, final_text: str) -> None:
        if self._hb_task: self._hb_task.cancel()
        if not self.msg_id or not self.chat_id: return
        try:
            import aiohttp
            from state import TG_TOKEN
            payload = {'chat_id': self.chat_id, 'message_id': self.msg_id,
                       'text': final_text, 'parse_mode': 'HTML'}
            async with aiohttp.ClientSession() as sess:
                await sess.post(f'https://api.telegram.org/bot{TG_TOKEN}/editMessageText', json=payload)
        except Exception:
            pass
    async def cancel(self) -> None:
        self.cancelled = True; self.phase = 'Cancelling...'
        if self._hb_task: self._hb_task.cancel()
        await self._edit(force=True)


_bt_progress: BtProgress | None = None


# ── Utility ──
def _utc_to_dam(dt) -> datetime:
    if isinstance(dt, pd.Timestamp): dt = dt.to_pydatetime()
    if dt.tzinfo is None: dt = dt.replace(tzinfo=timezone.utc)
    return dt + DAM_OFF


# ── Concurrent Analysis Sheets ──
def _add_concurrent_analysis_sheets(wb, trade_logs: list, pnl_key: str,
                                     outcome_key: str, slippage_key: str = None) -> None:
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    channel_sheets = {'touch': 'Touch_Trades', 'close': 'Close_Trades', 'hybrid': 'Hybrid_Trades'}
    channel_rows = {'touch': [], 'close': [], 'hybrid': []}
    for tr in trade_logs:
        ch = tr.get('trigger_type', 'touch')
        if ch in channel_rows:
            channel_rows[ch].append(tr)
    internal_keys = {'cycle_ts', 'cycle_time_str', 'cycle_close', 'trigger_type'}
    cols = [k for k in trade_logs[0].keys() if k not in internal_keys] if trade_logs else []
    made_sheets = []
    for ch, sheet_name in channel_sheets.items():
        ws = wb.create_sheet(sheet_name); made_sheets.append(ws)
        ws.append(cols)
        for cell in ws[1]: cell.fill = gray_fill; cell.font = Font(bold=True)
        for tr in channel_rows[ch]:
            ws.append([tr.get(c) for c in cols])
        for i in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20.0
    ws_cmp = wb.create_sheet("Performance_Comparison"); made_sheets.append(ws_cmp)
    ws_cmp.append(["Metric", "Touch (لمس مباشر)", "Close (إغلاق شمعة)", "Hybrid (هجين)"])
    for cell in ws_cmp[1]: cell.fill = gray_fill; cell.font = Font(bold=True)

    def metrics_for(rows):
        wins = [r for r in rows if r.get(outcome_key) == 'WIN']
        losses = [r for r in rows if r.get(outcome_key) == 'LOSS']
        gross_profit = sum((r.get(pnl_key) or 0) for r in wins)
        gross_loss = sum((r.get(pnl_key) or 0) for r in losses)
        net = sum((r.get(pnl_key) or 0) for r in rows)
        wr = round(100 * len(wins) / max(1, len(wins) + len(losses)), 1)
        rows_sorted = sorted(rows, key=lambda r: r.get('cycle_ts', 0))
        eq = 0.0; peak = 0.0; mdd = 0.0
        for r in rows_sorted:
            eq += (r.get(pnl_key) or 0); peak = max(peak, eq); mdd = min(mdd, eq - peak)
        avg_slip = None
        if slippage_key:
            slips = [r.get(slippage_key) for r in rows if r.get(slippage_key) is not None]
            avg_slip = round(sum(slips) / len(slips), 2) if slips else None
        return dict(total=len(rows), win=len(wins), loss=len(losses), wr=wr,
                    gp=round(gross_profit, 2), gl=round(gross_loss, 2), net=round(net, 2),
                    mdd=round(mdd, 2), avg_slip=avg_slip)
    m = {ch: metrics_for(channel_rows[ch]) for ch in channel_rows}
    rows_spec = [
        ("Total Trades", 'total'), ("Winning Trades", 'win'), ("Losing Trades", 'loss'),
        ("Win Rate (%)", 'wr'), ("Gross Profit ($)", 'gp'), ("Gross Loss ($)", 'gl'),
        ("Net PnL ($)", 'net'), ("Max Drawdown ($)", 'mdd'),
    ]
    if slippage_key: rows_spec.append(("Average Slippage (Pips)", 'avg_slip'))
    for label, key in rows_spec:
        ws_cmp.append([label, m['touch'][key], m['close'][key], m['hybrid'][key]])
    for i in range(1, 5): ws_cmp.column_dimensions[get_column_letter(i)].width = 26.0
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    for ws in made_sheets:
        for row in ws.iter_rows():
            for cell in row: cell.border = thin_border; cell.alignment = center_align


# ── Idealized Backtest Engine ──
async def run_gann_backtest(start_dt: datetime, end_dt: datetime) -> None:
    """Zero-friction Gann backtest (original engine, untouched logic)."""
    global _bt_progress
    _reset_data_source_stats()
    bot_state['is_backtesting'] = True
    fname = f"GannBT_{datetime.now(timezone.utc).strftime('%H%M%S')}.xlsx"
    exec_mode = bot_state.get('gann_execution_mode', 'instant')

    active_symbols = [s for s, on in bot_state['active_symbols'].items() if on]
    if not active_symbols:
        bot_state['is_backtesting'] = False
        return

    from state import _PRESET_EXCLUDED_KEYS
    from telegram_ui import send_tg_document, send_tg_msg

    first_sym_state = bot_state['symbol_state'][active_symbols[0]]
    enabled_tfs = [tf for tf, on in first_sym_state['gann_monitor_tfs'].items() if on] or ['5m']
    flt_type = first_sym_state['trend_filter_type']
    ttf = first_sym_state['trend_timeframe']
    desc_ttf = ttf.upper()

    if first_sym_state['gann_entry_mode'] == 'touch_trend':
        if flt_type == 'vwap': desc_mode = f"Touch(VWAP{first_sym_state['trend_vwap_period']}_{desc_ttf})\n"
        elif flt_type == 'ema': desc_mode = f"Touch(EMA{first_sym_state['trend_ema_period']}_{desc_ttf})\n"
        else: desc_mode = f"Touch(VWAP+EMA_{desc_ttf})\n"
    else:
        desc_mode = "Pure Touch"
    desc_be = " | 🛡️ BE" if first_sym_state['break_even_enabled'] else ""
    zf = first_sym_state['gann_zone_filter']
    if zf == 'star': desc_star = "⭐ الأصلية"
    elif zf == 'star_fan': desc_star = "⭐🌀 الأصلية والمروحة"
    else: desc_star = "📋 الكل"
    desc_tfs = "+".join(enabled_tfs); syms_label = "+".join(active_symbols)

    prog = BtProgress(label=f"{syms_label} جان H1→[{desc_tfs}] | {desc_mode} | {desc_star}{desc_be}", active_tfs=['H1'])
    _bt_progress = prog
    await prog.start(bot_state['chat_id'])

    res = {'win': 0, 'loss': 0, 'be': 0, 'total_prof': 0.0, 'total_win_usd': 0.0,
           'total_loss_usd': 0.0, 'peak_equity': 0.0, 'max_dd': 0.0,
           'trade_logs': [], 'cycle_logs': []}
    _earliest_1m_seen = {}

    try:
        delta_hours = int((end_dt - start_dt).total_seconds() / 3600)
        all_signals = []; all_candles_events = []

        for symbol in active_symbols:
            sym_state = bot_state['symbol_state'][symbol]
            cycle_h = sym_state['gann_cycle_hours']; tpsl_mode = sym_state['gann_tpsl_mode']
            pv = SYMBOL_INFO[symbol]['pip_value']; lot = sym_state['lot_size']
            cs = SYMBOL_INFO[symbol]['contract_size']
            prec = SYMBOL_INFO[symbol]['prec']

            quote = symbol.split('_')[1] if '_' in symbol else 'USD'
            _QUOTE_RATES = {'USD': 1.0, 'JPY': 1/150.0, 'AUD': 0.66, 'NZD': 0.61,
                            'EUR': 1.08, 'GBP': 1.27, 'CAD': 0.73, 'CHF': 1.11}
            quote_conv = _QUOTE_RATES.get(quote)
            if quote_conv is None:
                c_log(f"WARNING: unknown quote currency '{quote}' in {symbol}")
                quote_conv = 1.0

            await prog.set_phase(f'جلب بيانات الترند ({desc_ttf})...')
            max_period = max(sym_state['trend_vwap_period'], sym_state['trend_ema_period'], 100)
            trend_count = (delta_hours * (2 if ttf == '30m' else 1)) + max_period + 10
            candles_trend = await _fetch_candles_hybrid(symbol, ttf, count=trend_count, end_time=end_dt)
            if not candles_trend: continue

            df_trend = pd.DataFrame(candles_trend)
            if flt_type == 'vwap':
                p_vwap = sym_state['trend_vwap_period']
                df_trend['Typical_Price'] = (df_trend['high'] + df_trend['low'] + df_trend['close']) / 3
                df_trend['VWAP'] = (df_trend['Typical_Price'] * df_trend['volume']).rolling(window=p_vwap).sum() / df_trend['volume'].rolling(window=p_vwap).sum()
            if flt_type == 'ema':
                p_ema = sym_state['trend_ema_period']
                df_trend['EMA'] = df_trend['close'].ewm(span=p_ema, adjust=False).mean()

            df_trend.set_index('time', inplace=True)
            if flt_type == 'vwap': df_trend['macro_trend_up'] = df_trend['close'] > df_trend['VWAP']
            elif flt_type == 'ema': df_trend['macro_trend_up'] = df_trend['close'] > df_trend['EMA']
            elif flt_type == 'both':
                c1_up = df_trend['close'] > df_trend['VWAP']; c2_up = df_trend['close'] > df_trend['EMA']
                c1_dn = df_trend['close'] < df_trend['VWAP']; c2_dn = df_trend['close'] < df_trend['EMA']
                df_trend['macro_trend_up'] = np.where(c1_up & c2_up, True, np.where(c1_dn & c2_dn, False, None))

            anchor_gran = bot_state.get('gann_anchor_tf', '1h')
            await prog.set_phase(f'جلب بيانات {_anchor_label()}...')
            candles_h1 = await _fetch_candles_hybrid(symbol, anchor_gran, count=(delta_hours // _anchor_hours()) + 10, end_time=end_dt)
            if not candles_h1: continue

            await prog.set_phase('جلب شموع الفريمات الصغيرة...')
            monitor_tfs_data = {}
            days_diff = (end_dt - start_dt).days or 1
            need_1m = days_diff * 24 * 60 + 300
            mc_1m = await _fetch_candles_hybrid(symbol, '1m', count=need_1m, end_time=end_dt)
            if mc_1m:
                _earliest_1m_seen[symbol] = min(c['time'] for c in mc_1m)
                for c in mc_1m:
                    all_candles_events.append({'time': c['time'], 'symbol': symbol,
                                                'high': float(c['high']), 'low': float(c['low']),
                                                'close': float(c['close']), 'tf': '1m_track'})

            for btf in enabled_tfs:
                bmin = int(''.join(filter(str.isdigit, btf)))
                if 'h' in btf: bmin *= 60
                need_m = days_diff * 24 * (60 // max(bmin, 1)) + 300
                mc = await _fetch_candles_hybrid(symbol, btf, count=need_m, end_time=end_dt)
                if mc:
                    monitor_tfs_data[btf] = sorted(mc, key=lambda c: c['time'])
                    for c in mc:
                        all_candles_events.append({'time': c['time'], 'symbol': symbol,
                                                    'high': float(c['high']), 'low': float(c['low']),
                                                    'close': float(c['close']), 'tf': btf})

            start_ts = start_dt.timestamp(); end_ts = end_dt.timestamp()
            valid_h1 = [c for c in candles_h1 if start_ts <= (c['time'].timestamp() + 3600) <= end_ts]
            trend_freq = '30min' if ttf == '30m' else '1h'
            cycle_defs = _build_gann_cycle_defs(sym_state, valid_h1, mc_1m)

            for idx, cdef in enumerate(cycle_defs):
                if prog.cancelled: return
                await asyncio.sleep(0)
                t_start = cdef['t_start']; t_end = cdef['t_end']; close = cdef['close']
                levels = gann_calc_levels(symbol, close)
                f_mode = sym_state['gann_zone_filter']
                active_lv = [l for l in levels if l['dir'] != 'ref' and (
                    f_mode == 'all' or (f_mode == 'star' and l['star']) or
                    (f_mode == 'star_fan' and (l['star'] or l['fan'])))]
                res['cycle_logs'].append({'symbol': symbol, 'time_ts': t_start.timestamp(),
                                           'time_dt': t_start, 'close': close, 'levels': len(active_lv),
                                           'level_details': [{'key': l['key'], 'price': l['price'],
                                                               'dir': l['dir'], 'star': l['star'],
                                                               'fan': l['fan']} for l in active_lv]})
                level_used = set()
                exec_mode_bt = bot_state.get('gann_execution_mode', 'instant')
                spike_limit = bot_state.get('gann_spike_limit_pts', 20) * pv

                for btf, candles_m in monitor_tfs_data.items():
                    m_window = [c for c in candles_m if t_start <= c['time'] < t_end]
                    m_before = [c for c in candles_m if c['time'] < t_start]
                    atr_val = _gann_atr(m_before, sym_state['gann_atr_period']) if tpsl_mode == 'atr' else None
                    for bar in m_window:
                        bar_close = float(bar['close']); bar_time = bar['time']
                        bar_high = float(bar['high']); bar_low = float(bar['low'])
                        trend_up = True
                        if sym_state['gann_entry_mode'] == 'touch_trend':
                            # CAUSALITY FIX (no lookahead bias): the trend candle whose
                            # timestamp equals the floor of bar_time is STILL FORMING while
                            # bar_time is inside it (e.g. an M1 bar at 15:05 falls inside the
                            # 15:00 H1 candle, which only finishes closing at 15:59). Using
                            # that candle's close/EMA at 15:05 means judging the trend with a
                            # close price that hasn't happened yet. The live EA can only ever
                            # see the last *completed* candle, so we must step back one full
                            # trend interval to reference the previous, already-closed bar.
                            trend_time = bar_time.floor(trend_freq) - pd.Timedelta(trend_freq)
                            if trend_time in df_trend.index:
                                val = df_trend.loc[trend_time, 'macro_trend_up']
                                if isinstance(val, pd.Series): val = val.iloc[-1]
                                macro_trend_up = None if pd.isna(val) else bool(val)
                            else: macro_trend_up = None
                            if macro_trend_up is None: continue
                            trend_up = macro_trend_up

                        if bot_state.get('prot_cycle_inval', True):
                            inval_pts = bot_state.get('prot_cycle_inval_pts', 200) * pv
                            if abs(bar_close - close) > inval_pts:
                                active_lv = []; break

                        for lv in active_lv:
                            k = lv['key']; dir = lv['dir']; is_buy = (dir == 'dn')
                            if sym_state['gann_entry_mode'] == 'touch_trend':
                                if is_buy and not trend_up: continue
                                if not is_buy and trend_up: continue

                            channels = ['touch', 'close', 'hybrid'] if exec_mode_bt == 'all_concurrent' else [
                                'close' if exec_mode_bt == 'close' else 'hybrid' if exec_mode_bt == 'hybrid' else 'touch']
                            for channel in channels:
                                base_combo = f'{k}_{btf}' if bot_state.get('prot_allow_multi_tf', True) else k
                                combo_key = f"{base_combo}_{channel}" if exec_mode_bt == 'all_concurrent' else base_combo
                                if combo_key in level_used: continue
                                if channel == 'close':
                                    if abs(bar_close - lv['price']) != 0: continue
                                elif channel == 'hybrid':
                                    if not (bar_low <= lv['price'] <= bar_high): continue
                                    if bot_state.get('prot_spike_filter', True) and abs(bar_close - lv['price']) > spike_limit: continue
                                else:
                                    if not (bar_low <= lv['price'] <= bar_high): continue

                                entry = lv['price']
                                tf_tp = _gann_tf_tp(symbol, btf); tf_sl = _gann_tf_sl(symbol, btf)
                                if tpsl_mode == 'atr' and atr_val:
                                    sl_d = atr_val * sym_state['gann_atr_sl_mult']
                                    tp_d = atr_val * sym_state['gann_atr_tp_mult']
                                else:
                                    sl_d = tf_sl * pv; tp_d = tf_tp * pv
                                tp_px = entry + tp_d if is_buy else entry - tp_d
                                sl_px = entry - sl_d if is_buy else entry + sl_d
                                all_signals.append({
                                    'time': bar_time, 'symbol': symbol, 'is_buy': is_buy,
                                    'entry': entry, 'tp_px': tp_px, 'sl_px': sl_px,
                                    'sl_d': sl_d, 'tp_d': tp_d,
                                    'be_trigger_px': 'dynamic' if sym_state['break_even_enabled'] else None,
                                    'lot': lot, 'cs': cs, 'quote_conv': quote_conv,
                                    'tf': btf, 'combo_key': combo_key,
                                    'cycle_time': t_start, 'cycle_close': close,
                                    'level_key': k, 'trigger_type': channel,
                                })
                                level_used.add(combo_key)

        # PHASE 2: Event-driven simulation
        await prog.set_phase('محاكاة الصفقات الزمنية...')
        all_signals.sort(key=lambda x: x['time'])
        all_candles_events.sort(key=lambda x: x['time'])
        open_trades = []; closed_trades = []
        suspended_days = {}; suspend_trigger_time = {}
        daily_pl = 0.0; current_day = None; latest_price = {}
        signal_idx = 0; total_signals = len(all_signals)
        dd_limit = -float(bot_state['prot_daily_dd_usd'])
        profit_limit = float(bot_state['prot_daily_profit_usd'])
        total_events = len(all_candles_events)
        await prog.set_tf('محاكاة عائمة', total_events)

        for i, event in enumerate(all_candles_events):
            if i % 5000 == 0: await asyncio.sleep(0)
            if prog.cancelled: break
            t = event['time']; sym = event['symbol']
            h = event['high']; l = event['low']; c = event['close']
            day_str = _utc_to_dam(t).strftime('%Y-%m-%d')
            latest_price[sym] = c
            if day_str != current_day: current_day = day_str; daily_pl = 0.0

            if current_day not in suspended_days:
                floating_pl = 0.0; floating_pl_worst = 0.0
                for tr in open_trades:
                    lp = latest_price.get(tr['symbol'], tr['entry'])
                    diff = (lp - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp)
                    floating_pl += round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                    worst_px = (l if tr['is_buy'] else h) if tr['symbol'] == sym else lp
                    diff_worst = (worst_px - tr['entry']) if tr['is_buy'] else (tr['entry'] - worst_px)
                    floating_pl_worst += round(diff_worst * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                total_daily = daily_pl + floating_pl
                total_daily_worst = daily_pl + floating_pl_worst
                if dd_limit < 0 and total_daily_worst <= dd_limit:
                    suspended_days[current_day] = f'🛑 تراجع عائم (الحد {dd_limit}$)'
                elif profit_limit > 0 and total_daily >= profit_limit:
                    suspended_days[current_day] = f'✅ هدف عائم (الحد {profit_limit}$)'
                if current_day in suspended_days and current_day not in suspend_trigger_time:
                    suspend_trigger_time[current_day] = t
                if current_day in suspended_days:
                    for tr in open_trades:
                        lp = (l if tr['is_buy'] else h) if tr['symbol'] == sym else latest_price.get(tr['symbol'], tr['entry'])
                        diff = (lp - tr['entry']) if tr['is_buy'] else (tr['entry'] - lp)
                        p_usd = round(diff * tr['lot'] * tr['cs'] * tr['quote_conv'], 2)
                        tr['outcome'] = 'DAILY_LIMIT'; tr['p_usd'] = p_usd
                        tr['close_time'] = t; closed_trades.append(tr); daily_pl += p_usd
                    open_trades.clear()

            if current_day not in suspended_days:
                surviving = []
                for tr in open_trades:
                    if tr['symbol'] != sym: surviving.append(tr); continue
                    is_buy = tr['is_buy']; sl_current = tr['sl_current']; entry = tr['entry']
                    tp_px = tr['tp_px']; sl_d = tr['sl_d']; lot = tr['lot']
                    cs = tr['cs']; quote_conv = tr['quote_conv']
                    closed = False
                    pv_sym = SYMBOL_INFO[sym]['pip_value']
                    be_pts = bot_state['symbol_state'][sym].get('gann_be_trigger_points', 40)
                    atr_per = bot_state['symbol_state'][sym].get('gann_atr_period', 14)
                    cost_be_val = bot_state.get('prot_cost_be', True)

                    if not tr['be_activated'] and tr['be_trigger_px'] is not None:
                        test_px = h if is_buy else l
                        net_be = core_eval_break_even(is_buy, entry, test_px, pv_sym, be_pts, atr_per, cost_be_val)
                        if net_be is not None: tr['sl_current'] = net_be; tr['be_activated'] = True
                    _be_thresh = pv_sym * 2
                    if is_buy:
                        if l <= sl_current:
                            tr['outcome'] = 'BREAK_EVEN' if sl_current > entry - _be_thresh else 'LOSS'
                            tr['p_usd'] = round(abs(sl_current - entry) * lot * cs * quote_conv, 2) if tr['outcome'] == 'BREAK_EVEN' else -round(sl_d * lot * cs * quote_conv, 2)
                            closed = True
                        elif h >= tp_px:
                            tr['outcome'] = 'WIN'
                            tr['p_usd'] = round(tr['tp_d'] * lot * cs * quote_conv, 2); closed = True
                    else:
                        if h >= sl_current:
                            tr['outcome'] = 'BREAK_EVEN' if sl_current < entry + _be_thresh else 'LOSS'
                            tr['p_usd'] = round(abs(entry - sl_current) * lot * cs * quote_conv, 2) if tr['outcome'] == 'BREAK_EVEN' else -round(sl_d * lot * cs * quote_conv, 2)
                            closed = True
                        elif l <= tp_px:
                            tr['outcome'] = 'WIN'
                            tr['p_usd'] = round(tr['tp_d'] * lot * cs * quote_conv, 2); closed = True
                    if closed: tr['close_time'] = t; daily_pl += tr['p_usd']; closed_trades.append(tr)
                    else: surviving.append(tr)
                open_trades = surviving

            while signal_idx < total_signals and all_signals[signal_idx]['time'] <= t:
                sig = all_signals[signal_idx]; signal_idx += 1
                if current_day in suspended_days: continue
                if bot_state.get('prot_dam_time_filter', True):
                    sig_dam_time = (sig['time'] + timedelta(hours=3)).time()
                    if any(start <= sig_dam_time < end for start, end in _DAM_RESTRICTED_WINDOWS):
                        continue
                max_concurrent_bt = max(1, int(bot_state.get('prot_max_concurrent_trades', 4)))
                open_count_bt = sum(1 for tr in open_trades if tr['symbol'] == sig['symbol'])
                if open_count_bt >= max_concurrent_bt: continue
                sig['sl_current'] = sig['sl_px']; sig['be_activated'] = False
                open_trades.append(sig)
            await prog.tick(i, res['win'], res['loss'], res['be'], res['total_prof'])

        # Post-process
        for tr in closed_trades:
            if tr['outcome'] == 'WIN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] > 0):
                res['win'] += 1; res['total_win_usd'] += tr['p_usd']
            elif tr['outcome'] == 'LOSS' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] < 0):
                res['loss'] += 1; res['total_loss_usd'] += abs(tr['p_usd'])
            elif tr['outcome'] == 'BREAK_EVEN' or (tr['outcome'] == 'DAILY_LIMIT' and tr['p_usd'] == 0):
                res['be'] += 1
            res['total_prof'] += tr['p_usd']
            dir_str = 'BUY 📈' if tr['is_buy'] else 'SELL 📉'
            res['trade_logs'].append({
                'الزوج': tr['symbol'],
                'وقت الصفقة (DAM)': _utc_to_dam(tr['time']).strftime('%Y-%m-%d %H:%M'),
                'TF': tr['tf'], 'اتجاه': dir_str,
                'المستوى (الدخول)': f"{tr['entry']:.2f} ({tr['level_key']})",
                'الهدف (TP)': round(tr['tp_px'], 2),
                'الوقف (SL)': round(tr['sl_px'], 2),
                'النتيجة': tr['outcome'], 'ربح ($)': tr['p_usd'],
                'cycle_ts': tr['cycle_time'].timestamp(),
                'cycle_time_str': _utc_to_dam(tr['cycle_time']).strftime('%Y-%m-%d %H:%M'),
                'cycle_close': tr['cycle_close'],
                'trigger_type': tr.get('trigger_type', 'touch'),
            })
        res['trade_logs'].sort(key=lambda x: x['وقت الصفقة (DAM)'])
        running_eq = 5000.0; peak_eq = 5000.0; max_dd = 0.0
        for t_log in res['trade_logs']:
            running_eq += t_log['ربح ($)']; t_log['رصيد تراكمي ($)'] = round(running_eq, 2)
            if running_eq > peak_eq: peak_eq = running_eq
            dd = peak_eq - running_eq
            if dd > max_dd: max_dd = dd
        res['peak_equity'] = peak_eq; res['max_dd'] = max_dd

        if not res['trade_logs']:
            await prog.done('<b>باكتيست اكتمل ✅</b>\nلا توجد صفقات في هذا النطاق.')
            bot_state['is_backtesting'] = False; return

        await prog.set_phase('إنشاء ملف Excel...')
        sum_text = (
            f"<b>باكتيست جان اكتمل ✅</b>\n{syms_label} H1→[{desc_tfs}] | {desc_mode} | {desc_star}{desc_be}\n"
            f"{start_dt.strftime('%Y-%m-%d')} → {end_dt.strftime('%Y-%m-%d')}\n\n"
            f"Net: {'PROFIT ▲' if res['total_prof']>=0 else 'LOSS ▼'} ${round(res['total_prof'], 2)}\n"
            f"Win: +${round(res['total_win_usd'], 2)} ({res['win']})\n"
            f"Loss: -${round(res['total_loss_usd'], 2)} ({res['loss']})\n"
            f"Break-Even: $0 ({res['be']})\n"
            f"WR: {round(res['win']/max(1, res['win']+res['loss'])*100)}% ({len(res['trade_logs'])} صفقة)\n"
            f"Max DD: ${round(res['max_dd'],2)}\n"
            f"دورات H1: {len(res['cycle_logs'])} | Lot: {lot}\n\n"
            f"{_format_data_source_report()}"
        )
        # Excel generation
        wb = openpyxl.Workbook(); ws_trades = wb.active; ws_trades.title = "الصفقات"
        headers = ["الزوج", "وقت الصفقة (DAM)", "TF", "اتجاه", "المستوى (الدخول)",
                   "الهدف (TP)", "الوقف (SL)", "النتيجة", "ربح ($)", "رصيد تراكمي ($)"]
        ws_trades.append(headers)
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        be_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        for cell in ws_trades[1]: cell.fill = gray_fill; cell.font = Font(bold=True)
        for tr in res['trade_logs']:
            _OUTCOME_DISPLAY = {'WIN': 'WIN ✅', 'LOSS': 'LOSS ❌', 'BREAK_EVEN': 'BREAK_EVEN ⚖️', 'DAILY_LIMIT': 'DAILY_LIMIT ⏹️'}
            row_data = [tr['الزوج'], tr['وقت الصفقة (DAM)'], tr['TF'], tr['اتجاه'],
                        tr['المستوى (الدخول)'], tr['الهدف (TP)'], tr['الوقف (SL)'],
                        _OUTCOME_DISPLAY.get(tr['النتيجة'], tr['النتيجة']), tr['ربح ($)'], tr['رصيد تراكمي ($)']]
            ws_trades.append(row_data)
            fill = green_fill if tr['النتيجة'] == 'WIN' else red_fill if tr['النتيجة'] == 'LOSS' else be_fill if tr['النتيجة'] == 'BREAK_EVEN' else None
            if fill:
                for col in range(1, 11): ws_trades.cell(row=ws_trades.max_row, column=col).fill = fill
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        for row in ws_trades.iter_rows():
            for cell in row: cell.border = thin_border; cell.alignment = center_align
        for i in range(1, 11): ws_trades.column_dimensions[get_column_letter(i)].width = 22.0
        if exec_mode == 'all_concurrent':
            _add_concurrent_analysis_sheets(wb, res['trade_logs'], pnl_key='ربح ($)', outcome_key='النتيجة')

        # ── "دورات H1" sheet: every anchor cycle's close + the full level
        # ladder that resulted from it, listed directly underneath. This is
        # the data that was already being computed (cycle_logs) but never
        # actually written to the report -- only its COUNT was mentioned in
        # the Telegram summary line above ("دورات H1: N").
        if res['cycle_logs']:
            ws_cyc = wb.create_sheet("دورات H1")
            ws_cyc.append(["الزوج", "الدورة (DAM)", "إغلاق 1H", "عدد الصفقات", "ملاحظة"])
            for cell in ws_cyc[1]: cell.fill = gray_fill; cell.font = Font(bold=True)
            level_hdr_row_style = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
            for cdef_log in sorted(res['cycle_logs'], key=lambda c: c['time_ts']):
                trades_in_cycle = sum(1 for tr in res['trade_logs'] if tr.get('cycle_ts') == cdef_log['time_ts'])
                note = f"تم تنفيذ {trades_in_cycle} صفقة" if trades_in_cycle > 0 else "لم يلمس السعر أي مستوى"
                dam_time = _utc_to_dam(cdef_log['time_dt']).strftime('%Y-%m-%d %H:%M')
                ws_cyc.append([cdef_log['symbol'], dam_time, round(cdef_log['close'], 3),
                               trades_in_cycle, note])
                for cell in ws_cyc[ws_cyc.max_row]: cell.fill = gray_fill; cell.font = Font(bold=True)

                # every resulting level for this cycle, directly below its
                # summary row -- mirrors the same addition made to the MQL5
                # port's Excel export (anchor close + full level ladder).
                for lv in cdef_log.get('level_details', []):
                    tag = '⭐' if lv['star'] else ('🌀' if lv['fan'] else '•')
                    dir_lbl = 'مقاومة/بيع' if lv['dir'] == 'up' else 'دعم/شراء'
                    ws_cyc.append(["", "", f"{tag} {lv['key']} — {round(lv['price'], 3)}", "", dir_lbl])
                ws_cyc.append([])  # blank spacer row between cycles
            for i in range(1, 6): ws_cyc.column_dimensions[get_column_letter(i)].width = 24.0
            for row in ws_cyc.iter_rows():
                for cell in row: cell.border = thin_border; cell.alignment = center_align

        wb.save(fname)
        await prog.done(f'<b>باكتيست جان اكتمل ✅</b>\n{syms_label} — {len(res["trade_logs"])} صفقة\nجاري إرسال التقرير...')
        await send_tg_document(fname, sum_text)
        if os.path.exists(fname): os.remove(fname)
    except Exception as e:
        c_log(f'BT Error: {e}'); bot_state['is_backtesting'] = False
        if _bt_progress:
            import html
            try: await _bt_progress.done(f'❌ خطأ داخلي:\n{html.escape(str(e))}')
            except Exception: pass
    finally:
        bot_state['is_backtesting'] = False


